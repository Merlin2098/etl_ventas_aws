from __future__ import annotations

import io
from typing import Iterator

from openpyxl import load_workbook

from src.lambda_ingestion.common.errors import FileParseError
from src.lambda_ingestion.common.parser_base import SalesParser


class ExcelSalesParser(SalesParser):
    division = "supermercado"

    def parse(self, raw_bytes: bytes) -> Iterator[dict]:
        try:
            workbook = load_workbook(
                io.BytesIO(raw_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            raise FileParseError(
                stage="parse", cause=f"unreadable_xlsx: {exc}"
            ) from exc

        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise FileParseError(stage="parse", cause="empty_xlsx") from exc

        fieldnames = [str(cell) if cell is not None else "" for cell in header]

        yielded_any = False
        for row in rows_iter:
            if row is None or all(cell is None for cell in row):
                continue
            yielded_any = True
            yield dict(zip(fieldnames, row))

        if not yielded_any:
            raise FileParseError(stage="parse", cause="no_rows_in_xlsx")
